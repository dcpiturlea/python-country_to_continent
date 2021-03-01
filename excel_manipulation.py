import openpyxl
import easygui
import country_manipulation
#import os

def open_xl(path):
    wb = openpyxl.load_workbook(path)
    return wb


#deschidere sheet "sheet1"
#loop prin toate linii, coloana 1
#chemata functia care returneaza zona pentru tara aflata pe coloana 1
#scrierea zonei pe aceeasi linie, coloana 2

def get_all_countries(wb_status, filepath):
    worksheet_names = wb_status.sheetnames
    sheet_index = worksheet_names.index("Sheet1")
    wb_status.active = sheet_index
    sheet_keyword = wb_status.active

    #last_column = sheet_keyword.max_column
    last_row = sheet_keyword.max_row

    for i in range(2, last_row):
        country = sheet_keyword.cell(row=i, column=1).value
        try:
            zone = country_manipulation.country_to_continent(country)
        except:
            zone = "N/A"
        if str(zone) != "Europe":
            if str(zone) != "N/A":
                zone = "Others"
        sheet_keyword.cell(row=i, column=2).value = zone
        print("row: " + str(i) + " zone: " + zone)
    wb.save(filepath)


path = easygui.fileopenbox()
#filepath = os.path.basename(path)
wb = open_xl(path)
get_all_countries(wb, path)
